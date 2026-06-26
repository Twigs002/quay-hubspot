"""
Parse the "Brokering Divisions" sheet of Divisions Area Breakdown.xlsx
into a clean JSON file the dashboard can consume.

Source layout (sheet "Brokering Divisions"):
  Header row:
    TEAM | BROKERS 1 | BROKERS 2 | BROKERS 3 | PROPERTY SPECIALIST 1
        | PROPERTY SPECIALIST 2 | PROPERTY SPECIALIST 3
        | SUBURBS (As per CMA) | HUBSPOT EMAIL ADDRESS
        | HUBSPOT OWNER ID | HUBSPOT DIVISION | TYPE
  Then alternating:
    - Section header rows  → "RESI SALES - CITY BOWL & ATLANTIC SEABOARD"
      (all cols except A are None/empty)
    - 3-row team blocks    → name row, then 1-2 contact rows (emails +
      phones, sometimes interleaved)

Output:
  data/divisions.json
  {
    "generated": "<UTC iso>",
    "sections": [
      {
        "name": "...",
        "teams": [
          {
            "name": "Assassins",
            "type": "Sales",
            "suburbs": "Vredehoek ST, Oranjezicht ST, ...",
            "hubspot_email": "assassins@greeffcity.co.za",
            "hubspot_owner_id": "61949424",
            "hubspot_division": "Assassins",
            "brokers":     [{ "name": "...", "email": "...", "phone": "..." }, ...],
            "specialists": [{ "name": "...", "email": "...", "phone": "..." }, ...]
          }
        ]
      }
    ]
  }
"""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUT  = ROOT / "data" / "divisions.json"

# The xlsx is gitignored — keep it local. Look in data/ first, then fall
# back to the most recent ~/Downloads/Divisions* the user has, so the
# typical flow ("download, run script") works without a manual move.
def _resolve_source() -> Path:
    explicit = ROOT / "data" / "divisions_source.xlsx"
    if explicit.exists():
        return explicit
    downloads = Path.home() / "Downloads"
    candidates = sorted(
        downloads.glob("Divisions*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if candidates:
        return candidates[0]
    sys.exit(
        "No source xlsx found. Either:\n"
        "  - drop the file at data/divisions_source.xlsx, or\n"
        f"  - export Divisions Area Breakdown.xlsx into {downloads}"
    )

SRC = _resolve_source()

EMAIL_RX = re.compile(r"^[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}$")
PHONE_RX = re.compile(r"[\d\s+()\-/]{7,}")

# Columns (1-indexed in xlsx, 0-indexed in tuples):
#   A=team, B=broker1, C=broker2, D=broker3,
#   E=spec1,  F=spec2,    G=spec3,
#   H=suburbs, I=hs_email, J=hs_owner_id, K=hs_division, L=type
BROKER_COLS = [1, 2, 3]   # 0-indexed
SPEC_COLS   = [4, 5, 6]


def _norm(s):
    if s is None:
        return None
    # HubSpot owner IDs come through as floats (61949424.0) because Excel
    # stores them numerically. Strip the trailing .0 so they round-trip
    # cleanly into HubSpot URLs.
    if isinstance(s, float) and s.is_integer():
        s = str(int(s))
    s = str(s).strip()
    return s or None


def _is_email(s):
    return bool(s) and bool(EMAIL_RX.match(s))


def _is_phone(s):
    if not s:
        return False
    # Strip non-digits to count digits — phones have ≥7 digits in SA.
    digits = re.sub(r"\D", "", s)
    return len(digits) >= 7 and not _is_email(s)


def _is_section_header(row):
    """A section header has text in col A and only None in B..G."""
    if not _norm(row[0]):
        return False
    rest = [_norm(row[i]) for i in range(1, 7)]
    return all(v is None for v in rest)


def _is_team_header(row):
    """A team header row has a name in A and at least one broker name in B-D."""
    if not _norm(row[0]):
        return False
    if _is_section_header(row):
        return False
    return any(_norm(row[c]) for c in BROKER_COLS + SPEC_COLS)


def _slot_for(name, col_idx):
    """Track which slot a contact belongs to so we can match emails/phones back."""
    return (name or "").lower(), col_idx


def parse():
    if not SRC.exists():
        sys.exit(f"Source not found: {SRC}")
    wb = openpyxl.load_workbook(SRC, data_only=True)
    if "Brokering Divisions" not in wb.sheetnames:
        sys.exit("Sheet 'Brokering Divisions' missing from workbook")
    ws = wb["Brokering Divisions"]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit("Empty sheet")

    # Find the header row (skip any pre-amble), then start parsing after it.
    header_idx = 0
    for i, r in enumerate(rows[:10]):
        if r and _norm(r[0]) == "TEAM":
            header_idx = i
            break

    sections = []
    current_section = None
    current_team = None
    contact_rows_consumed = 0

    def _flush_team():
        nonlocal current_team
        if current_team and current_section is not None:
            current_section["teams"].append(current_team)
        current_team = None

    def _new_section(name):
        nonlocal current_section
        if current_section and current_section["teams"]:
            sections.append(current_section)
        current_section = {"name": name, "teams": []}

    # Default catch-all for any teams that appear before the first section
    # header (none in the current file, but defensive).
    _new_section("Unsorted")

    for row in rows[header_idx + 1:]:
        if not row:
            continue
        if all(_norm(c) is None for c in row):
            # Blank row — end of any in-progress team.
            _flush_team()
            contact_rows_consumed = 0
            continue

        if _is_section_header(row):
            _flush_team()
            contact_rows_consumed = 0
            _new_section(_norm(row[0]))
            continue

        if _is_team_header(row):
            _flush_team()
            contact_rows_consumed = 0
            team_name = _norm(row[0])
            brokers = [
                {"name": _norm(row[c]), "email": None, "phone": None}
                for c in BROKER_COLS if _norm(row[c])
            ]
            specs = [
                {"name": _norm(row[c]), "email": None, "phone": None}
                for c in SPEC_COLS if _norm(row[c])
            ]
            current_team = {
                "name":             team_name,
                "type":             _norm(row[11]) or "",
                "suburbs":          _norm(row[7]) or "",
                "hubspot_email":    _norm(row[8]) or "",
                "hubspot_owner_id": _norm(row[9]) or "",
                "hubspot_division": _norm(row[10]) or "",
                "brokers":          brokers,
                "specialists":      specs,
            }
            continue

        # Otherwise: contact row beneath a team header. Two rows max
        # (emails, phones — sometimes interleaved). Walk B-D for brokers
        # and E-G for specialists, slot value into the matching contact.
        if not current_team or contact_rows_consumed >= 2:
            continue

        contacts = current_team["brokers"] + [None] + current_team["specialists"]
        # Layout the row values against the contacts (col B → brokers[0],
        # col C → brokers[1], etc.).
        slot_pairs = (
            list(zip(BROKER_COLS, current_team["brokers"]))
            + list(zip(SPEC_COLS, current_team["specialists"]))
        )
        for col_idx, contact in slot_pairs:
            val = _norm(row[col_idx])
            if not val:
                continue
            if _is_email(val) and not contact["email"]:
                contact["email"] = val
            elif _is_phone(val) and not contact["phone"]:
                contact["phone"] = val
            elif not contact["email"] and "@" in val:
                # Looser email check (some rows have surrounding whitespace).
                contact["email"] = val
            elif not contact["phone"]:
                contact["phone"] = val
        contact_rows_consumed += 1

    _flush_team()
    if current_section and current_section["teams"]:
        sections.append(current_section)

    # Drop the "Unsorted" catch-all if it ended up empty.
    sections = [s for s in sections if s["teams"]]

    # Stats — useful for sanity-checking the parse.
    team_count    = sum(len(s["teams"]) for s in sections)
    broker_count  = sum(len(t["brokers"])      for s in sections for t in s["teams"])
    spec_count    = sum(len(t["specialists"])  for s in sections for t in s["teams"])
    with_hubspot  = sum(1 for s in sections for t in s["teams"] if t["hubspot_owner_id"])

    out = {
        "generated": datetime.now(timezone.utc).isoformat(),
        "source":    "scripts/parse_divisions.py · data/divisions_source.xlsx",
        "stats": {
            "sections":           len(sections),
            "teams":              team_count,
            "brokers":            broker_count,
            "specialists":        spec_count,
            "teams_with_hubspot": with_hubspot,
        },
        "sections": sections,
    }
    OUT.write_text(json.dumps(out, indent=2, ensure_ascii=False))
    print(f"Wrote {OUT.relative_to(ROOT)}")
    print(f"  {len(sections)} sections · {team_count} teams "
          f"· {broker_count} brokers · {spec_count} specialists "
          f"· {with_hubspot} teams with HubSpot owner ID")


if __name__ == "__main__":
    parse()
